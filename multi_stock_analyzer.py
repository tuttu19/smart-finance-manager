import pandas as pd
import yfinance as yf
import ta
import openpyxl
import numpy as np
from sklearn.linear_model import LinearRegression
from datetime import timedelta, date
from openpyxl.chart import LineChart, Reference, PieChart, BarChart, Series
from openpyxl.chart.label import DataLabelList
import os
import requests

# === Setup Paths ===
input_file = "Stock_Analyzer.xlsx"
output_file = "Stock_Analyzer_Multi_Output.xlsx"

# === Load Workbook and Data ===
df = pd.read_excel(input_file)
df.columns = df.columns.str.strip()
wb = openpyxl.load_workbook(input_file)
ws_main = wb.active

# === Load Settings ===
ws_settings = wb["Settings"]
settings = {ws_settings[f"A{i}"].value: ws_settings[f"B{i}"].value for i in range(2, 20)}

base_currency = settings.get("Base Currency", "USD")
currency_symbol = settings.get("Currency Symbol", "$")
forecast_days = int(settings.get("Forecast Days", 7))
min_data_points = int(settings.get("Minimum Data Points", 10))
rsi_alert_threshold = float(settings.get("RSI Alert Threshold", 30))
profit_alert_minimum = float(settings.get("Profit Alert Minimum", 50))
profit_target_threshold = float(settings.get("Profit Target Threshold", 100))
use_live = settings.get("Use Live Price", "No").strip().lower() == "yes"
capital_api_key = settings.get("Capital API Key", "")

# === Live Exchange Rates ===
def get_live_exchange_rate(base, target):
    try:
        url = f"https://api.exchangerate.host/latest?base={base}&symbols={target}"
        res = requests.get(url)
        if res.status_code == 200:
            return res.json()["rates"][target]
    except:
        pass
    return 1

currency_rate = get_live_exchange_rate("USD", base_currency)

# === Portfolio Summary ===
portfolio_investment = 0
portfolio_profit = 0
rsi_values = []
count_buy = 0
count_hold = 0
count_sell = 0
count_alerts = 0
symbols = []
profits = []
advice_labels = []

# === Helper: Get live price from Capital.com ===
def get_live_price(symbol):
    try:
        auth_url = "https://api-capital.backend-capital.com/api/v1/session"
        headers = {
            "X-CAP-API-KEY": capital_api_key,
            "Content-Type": "application/json"
        }
        auth_response = requests.post(auth_url, headers=headers, json={})
        if auth_response.status_code == 200:
            cst = auth_response.headers.get("CST")
            token = auth_response.headers.get("X-SECURITY-TOKEN")
            price_url = f"https://api-capital.backend-capital.com/api/v1/prices/{symbol}"
            price_headers = {
                "X-CAP-API-KEY": capital_api_key,
                "CST": cst,
                "X-SECURITY-TOKEN": token
            }
            response = requests.get(price_url, headers=price_headers)
            if response.status_code == 200:
                data = response.json()
                return data['prices'][-1]['closePrice']['bid']
    except:
        pass
    return None

# === Process Each Row ===
for idx, row in df.iterrows():
    if pd.isna(row['Stock Symbol']) or pd.isna(row['Days of Data']) or pd.isna(row['Investment (USD)']):
        print(f"⚠️ Skipping row {idx + 2} due to missing data.")
        continue

    symbol = row['Stock Symbol']
    days = int(row['Days of Data'])
    amount = float(row['Investment (USD)'])
    print(f"\n🔄 Processing: {symbol}...")

    data = yf.download(symbol, period=f"{days}d", interval="1d", auto_adjust=True).dropna()
    if data.empty or len(data) < min_data_points:
        print(f"⚠️ Not enough data for {symbol}. Skipping.")
        continue

    close_series = pd.Series(data['Close'].values.ravel(), index=data.index)
    rsi_calc = ta.momentum.RSIIndicator(close=close_series)
    data['RSI'] = rsi_calc.rsi()
    latest_rsi = data['RSI'].dropna().iloc[-1]

    advice = "Buy" if latest_rsi < 30 else "Sell" if latest_rsi > 70 else "Hold"

    if use_live:
        live_price = get_live_price(symbol)
        if live_price:
            latest_price_usd = float(live_price)
            data_source = "Capital.com"
        else:
            latest_price_usd = float(data['Close'].iloc[-1])
            data_source = "Yahoo (fallback)"
    else:
        latest_price_usd = float(data['Close'].iloc[-1])
        data_source = "Yahoo"

    latest_price = latest_price_usd * currency_rate
    shares = int(amount // latest_price_usd)

    data = data.reset_index()
    data['DateOrdinal'] = pd.to_datetime(data['Date']).map(pd.Timestamp.toordinal)
    model = LinearRegression()
    model.fit(data[['DateOrdinal']], data['Close'])

    last_date = date.today()
    future_dates = [last_date + timedelta(days=i) for i in range(1, forecast_days + 1)]
    future_ordinals = np.array([d.toordinal() for d in future_dates]).reshape(-1, 1)
    predicted_prices_usd = model.predict(future_ordinals).flatten()
    predicted_prices = predicted_prices_usd * currency_rate

    best_idx = np.argmax(predicted_prices)
    best_sell_date = future_dates[best_idx].strftime('%Y-%m-%d')
    best_sell_price = predicted_prices[best_idx].item()
    profit = (best_sell_price - latest_price) * shares

    if latest_rsi < 30 and profit > 0:
        live_advice = "🔥 Strong Buy – Oversold & Profit Likely"
    elif latest_rsi > 70 and profit > 0:
        live_advice = "🔺 Consider Selling – Overbought + Peak Soon"
    elif abs(latest_rsi - 50) <= 10 and profit <= 0:
        live_advice = "⚠️ Wait – Neutral RSI & Low Forecast"
    else:
        live_advice = f"📊 {advice} – Review forecast"

    if profit >= profit_target_threshold:
        alert = "🎯 Target Hit"
    elif latest_rsi < rsi_alert_threshold:
        alert = "⚠️ RSI Too Low"
    elif profit < profit_alert_minimum:
        alert = "💸 Low Profit"
    else:
        alert = "✅ OK"

    if advice == "Buy":
        count_buy += 1
    elif advice == "Hold":
        count_hold += 1
    elif advice == "Sell":
        count_sell += 1

    portfolio_investment += amount
    portfolio_profit += profit
    rsi_values.append(latest_rsi)
    symbols.append(symbol)
    profits.append(round(profit, 2))
    advice_labels.append(advice)

# === Clean Old Dashboard ===
if "Dashboard" in wb.sheetnames:
    del wb["Dashboard"]
ws_dash = wb.create_sheet("Dashboard")

# === Table Summary ===
ws_dash.append(["Stock Symbol", "Advice", "Profit", "RSI", "Alert"])
for i in range(len(symbols)):
    alert_flag = "Yes" if rsi_values[i] < rsi_alert_threshold or profits[i] < profit_alert_minimum or profits[i] >= profit_target_threshold else "No"
    ws_dash.append([symbols[i], advice_labels[i], profits[i], round(rsi_values[i], 2), alert_flag])

# === Bar Chart: Profit ===
bar_chart = BarChart()
bar_chart.title = "Profit by Stock"
bar_chart.y_axis.title = f"Profit ({currency_symbol})"
bar_chart.x_axis.title = "Stock Symbol"

data_ref = Reference(ws_dash, min_col=3, min_row=2, max_row=1 + len(symbols))
cats_ref = Reference(ws_dash, min_col=1, min_row=2, max_row=1 + len(symbols))
bar_chart.add_data(data_ref, titles_from_data=False)
bar_chart.set_categories(cats_ref)
bar_chart.height = 10
bar_chart.width = 20
ws_dash.add_chart(bar_chart, "G2")

# === Pie Chart: Advice Split ===
advice_summary = {"Buy": count_buy, "Hold": count_hold, "Sell": count_sell}
ws_dash.append([])
ws_dash.append(["Advice", "Count"])
for label, count in advice_summary.items():
    ws_dash.append([label, count])

pie_chart = PieChart()
pie_chart.title = "Advice Distribution"
data = Reference(ws_dash, min_col=2, min_row=len(symbols) + 4, max_row=len(symbols) + 6)
labels = Reference(ws_dash, min_col=1, min_row=len(symbols) + 4, max_row=len(symbols) + 6)
pie_chart.add_data(data, titles_from_data=False)
pie_chart.set_categories(labels)
pie_chart.dataLabels = DataLabelList()
pie_chart.dataLabels.showPercent = True
ws_dash.add_chart(pie_chart, "G20")

# === Save Workbook ===
wb.save(output_file)
print(f"\n📊 Dashboard updated and all results saved to: {output_file}")
