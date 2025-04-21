import pandas as pd
import yfinance as yf
import ta
import openpyxl
import numpy as np
from sklearn.linear_model import LinearRegression
from datetime import timedelta
from openpyxl.chart import LineChart, Reference

# === STEP 1: Load Excel ===
excel_data = pd.read_excel("Stock_Analyzer.xlsx")
excel_data.columns = excel_data.columns.str.strip()
print("Excel columns found:", excel_data.columns.tolist())

# === STEP 2: Read Inputs ===
symbol = excel_data.loc[0, 'Stock Symbol']
days = int(excel_data.loc[0, 'Days of Data'])
amount = float(excel_data.loc[0, 'Investment (USD)'])

# === STEP 3: Download Stock Data ===
data = yf.download(symbol, period=f'{days}d', interval='1d')
data = data.dropna()

# === STEP 4: RSI Calculation ===
close_series = data['Close'].squeeze()
rsi_calc = ta.momentum.RSIIndicator(close=close_series)
data['RSI'] = rsi_calc.rsi()
latest_rsi = data['RSI'].dropna().iloc[-1]

# === STEP 5: Advice Logic ===
if latest_rsi < 30:
    advice = "Buy"
elif latest_rsi > 70:
    advice = "Sell"
else:
    advice = "Hold"

# === STEP 6: Latest Price & Shares Calculation ===
closing_prices = data['Close'].dropna()
if closing_prices.empty:
    print("❌ No valid closing price data available.")
    exit()

latest_price = closing_prices.iloc[-1].item()
if pd.isna(latest_price):
    print("❌ Error: Latest price is NaN.")
    exit()

shares = int(amount // latest_price)

# === STEP 7: Forecasting with Linear Regression ===
data = data.reset_index()
data['DateOrdinal'] = pd.to_datetime(data['Date']).map(pd.Timestamp.toordinal)

X = data[['DateOrdinal']]
y = data['Close']
model = LinearRegression()
model.fit(X, y)

# Predict next 7 days
future_days = 7
last_date = data['Date'].iloc[-1]
future_dates = [last_date + timedelta(days=i) for i in range(1, future_days + 1)]
future_ordinals = np.array([date.toordinal() for date in future_dates]).reshape(-1, 1)
predicted_prices = model.predict(future_ordinals)

# Best day to sell
best_day_index = np.argmax(predicted_prices)
best_sell_date = future_dates[best_day_index].strftime('%Y-%m-%d')
best_sell_price = predicted_prices[best_day_index].item()
profit = (best_sell_price - latest_price) * shares

# === STEP 8: Write Summary to Excel ===
wb = openpyxl.load_workbook("Stock_Analyzer.xlsx")
ws = wb.active

ws["D1"] = "Latest RSI"
ws["D2"] = round(latest_rsi, 2)
ws["E1"] = "Advice"
ws["E2"] = advice
ws["F1"] = "Latest Price"
ws["F2"] = round(latest_price, 2)
ws["G1"] = "Buyable Shares"
ws["G2"] = shares
ws["H1"] = "Forecasted Best Sell Date"
ws["H2"] = best_sell_date
ws["I1"] = "Forecasted Sell Price"
ws["I2"] = round(best_sell_price, 2)
ws["J1"] = "Estimated Profit"
ws["J2"] = round(profit, 2)

# === STEP 9: Write Forecast Table for Chart ===
ws["L1"] = "Forecast Date"
ws["M1"] = "Predicted Price"
for i, (date, price) in enumerate(zip(future_dates, predicted_prices), start=2):
    ws[f"L{i}"] = date.strftime('%Y-%m-%d')
    ws[f"M{i}"] = round(price.item(), 2)

# === STEP 10: Write combined actual + forecast data for chart ===
ws["O1"] = "Chart Date"
ws["P1"] = "Actual Price"
ws["Q1"] = "Forecasted Price"

# Get last 30 actual dates and prices
actual_dates = data['Date'].dt.strftime('%Y-%m-%d').tolist()[-30:]
actual_prices = data['Close'].values[-30:].tolist()

# Combine both
combined_dates = actual_dates + [d.strftime('%Y-%m-%d') for d in future_dates]
combined_actual = actual_prices + [None] * len(future_dates)
combined_forecast = [None] * len(actual_prices) + [p.item() for p in predicted_prices]

for i, (d, a, f) in enumerate(zip(combined_dates, combined_actual, combined_forecast), start=2):
    ws[f"O{i}"] = d
    ws[f"P{i}"] = a[0] if isinstance(a, (list, np.ndarray)) else a
    ws[f"Q{i}"] = round(f, 2) if f else None

# === STEP 11: Create combined chart ===
chart = LineChart()
chart.title = "30-Day Price vs 7-Day Forecast"
chart.x_axis.title = "Date"
chart.y_axis.title = "Price (USD)"

actual_ref = Reference(ws, min_col=16, min_row=1, max_row=1 + len(combined_dates))   # P
forecast_ref = Reference(ws, min_col=17, min_row=1, max_row=1 + len(combined_dates)) # Q
dates_ref = Reference(ws, min_col=15, min_row=2, max_row=1 + len(combined_dates))    # O

chart.add_data(actual_ref, titles_from_data=True)
chart.add_data(forecast_ref, titles_from_data=True)
chart.set_categories(dates_ref)

chart.style = 2
chart.width = 20
chart.height = 10
ws.add_chart(chart, "S2")

# === STEP 12: Save ===
wb.save("Stock_Analyzer.xlsx")

# === STEP 13: Console Output ===
print(f"\n✅ {symbol.upper()} advice: {advice}")
print(f"📊 RSI: {latest_rsi:.2f}")
print(f"💵 Latest Price: ${latest_price:.2f}")
print(f"🧮 Buyable Shares with ${amount:.2f}: {shares}")
print(f"\n🔮 Forecasted Sell Price in next {future_days} days: ${best_sell_price:.2f}")
print(f"📅 Best day to sell: {best_sell_date}")
print(f"💰 Estimated profit: ${profit:.2f}")
